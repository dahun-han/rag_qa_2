"""export.py — QAItem -> QA/USED_DOCS 시트, retrieval/answer 요약, INPUT_DATA 엑셀 파생"""

from __future__ import annotations

import glob
import json
import os
import re
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd

from answer_gen import Nugget
from candidate_pool import CandidateChunk
from query_pipeline import ChunkRef
from retrieval_sim import RetrievalSim


def format_kr_date(d: datetime) -> str:
    """"2026. 8. 12"처럼 앞자리 0 없이 연.월.일을 점으로 구분해 표기한다."""
    return f"{d.year}. {d.month}. {d.day}"


@dataclass
class QAItem:
    index: str
    seq: int  # 청킹방식 내 순번(1부터, start_seq 반영됨) - item_id(ID_BASE + seq)의 근거
    chunking_method: str
    data_type: str
    difficulty: str
    persona: str
    category: str
    subcategory: str
    query: str
    positive_docs: dict
    candidate_pool: list[CandidateChunk]
    positive_chunk_ids: list[str]
    ground_truth: str
    nuggets: list[Nugget]
    reason: str
    positive_chunk_meta: dict = field(default_factory=dict)


QA_COLUMNS = [
    "index", "as_of_date", "chunking_method", "data_type", "difficulty", "persona", "category",
    "subcategory", "query", "ground_truth", "positive_docs",
    "positive_chunk_ids", "positive_chunk_meta",
    "negative_docs", "negative_chunk_ids", "negative_chunk_meta",
    "nuggets", "reason", "검수", "납품일자", "폐기 사유", "query_수정",
]
REVIEW_STATUS_DEFAULT = "대기"  # 검수 컬럼 기본값 - 사람이 검수 후 직접 값을 바꿔 넣는다
# QA 시트에 노출하는 negative는 candidate_pool 전체(문서 50개 단위로 문항당 최대 ~2,000개)가
# 아니라 hard 등급만, 최대 이 개수로 제한한다 — 전체를 다 넣으면 Excel 셀 문자수 제한
# (32,767자)을 넘길 수 있다.
NEGATIVE_CHUNK_DISPLAY_LIMIT = 5
CHUNK_COLUMNS = [
    "index", "chunking_method", "chunk_id", "doc_id(uuid)", "doc_nm",
    "category", "subcategory", "is_positive", "neg_grade", "content",
]
DOC_COLUMNS = ["index", "doc_id(uuid)", "doc_nm", "system_nm", "category", "subcategory", "is_positive_doc"]
INPUT_COLUMNS = [
    "chunk_id", "uuid", "doc_nm", "category", "subcategory", "content", "topic_list", "start_char", "end_char",
]

# 청킹방식 3개는 문항을 공유하지 않는다(방식마다 완전히 독립적으로 생성·번호 매김 -
# main.py의 최종 재정렬 참고). 그래서 ID도 방식별로 1000단위 대역을 나눠 "QA_" + 숫자
# 하나로 통일한다 - QA 시트의 index, retrieval/answer json의 item_id(=파일명)가 같은
# 문항에 대해 항상 같은 값을 가리켜, 어느 산출물에서 보든 그 값만으로 바로 대조할 수
# 있다. seq는 이번 실행에서 새로 매긴 1부터의 순번이고, 여기에 sampling_plan.yaml의
# start_seq(이미 다른 배치에서 만들어져 확인한 개수)를 더해 번호가 이어지게 한다
# (main.py에서 이미 더해진 채로 it.seq에 들어옴).
ID_BASE = {
    "recursive_300_100": 1000,
    "recursive_600_200": 2000,
    "recursive_1000_300": 3000,
}
ID_PREFIX = "QA_"

# retrieval/answer json의 difficulty는 QA 시트(하/중/상)와 달리 영문(easy/normal/hard)으로 낸다.
DIFFICULTY_EN = {"하": "easy", "중": "normal", "상": "hard"}


def item_id(chunking_method: str, seq: int) -> str:
    return f"{ID_PREFIX}{ID_BASE[chunking_method] + seq}"


_RUN_DIR_RE = re.compile(r"^\d{6}_\d{6}$")
_ITEM_ID_RE = re.compile(r"^QA_(\d+)\.json$")
_METHOD_BY_BASE = {base: method for method, base in ID_BASE.items()}


def resolve_start_seq(output_root: str = "output") -> dict[str, int]:
    """가장 최근 실행 폴더(output_root 아래 YYMMDD_HHMMSS 폴더 중 이름순 최댓값 -
    폴더명이 실행 시각이라 이름순 정렬이 곧 시간순이다)의 retrieval/answer json
    파일명(item_id)에서 청킹방식별 최대 seq를 역산한다 - main()이 이번 실행의
    번호를 그 다음부터 이어서 매기는 데 쓴다(§export.ID_BASE와 정반대 방향의 계산).

    old_test처럼 YYMMDD_HHMMSS 패턴이 아닌 폴더는 애초에 후보에서 걸러진다.
    가장 최근 폴더에 특정 방식의 json이 하나도 없으면(예: 그 방식은 비중 0이었거나
    실행이 산출물 없이 끝남) 그 방식은 결과 dict에 키 자체가 없다 - 호출부가
    .get(method, 0)으로 처음부터(0) 시작하게 둔다."""
    if not os.path.isdir(output_root):
        return {}
    run_dirs = [d for d in os.listdir(output_root) if _RUN_DIR_RE.match(d)]
    if not run_dirs:
        return {}
    latest = max(run_dirs)

    max_seq: dict[str, int] = {}
    for sub in ("retrieval", "answer"):
        for path in glob.glob(os.path.join(output_root, latest, sub, "QA_*.json")):
            m = _ITEM_ID_RE.match(os.path.basename(path))
            if not m:
                continue
            item_num = int(m.group(1))
            method = _METHOD_BY_BASE.get(item_num // 1000 * 1000)
            if method is None:
                continue
            seq = item_num - ID_BASE[method]
            max_seq[method] = max(max_seq.get(method, 0), seq)
    return max_seq


def _hard_negative_chunks(
    candidate_pool: list[CandidateChunk], data_type: str, limit: int = NEGATIVE_CHUNK_DISPLAY_LIMIT,
) -> list:
    """QA 시트에 노출할 negative 청크를 hard 등급만, candidate_pool에 이미 들어있는
    순서(문항마다 랜덤 셔플됨, build_candidate_pool 참고) 그대로 앞에서 최대 limit개
    골라 ChunkRef 리스트로 반환한다. data_type=all_pos는 정의상 "positive 청크만으로
    답이 되고 negative에 의존하지 않는 문항"이라 candidate_pool에 distractor가 섞여
    있어도(모의 검색 강건성 테스트용) 항상 빈 리스트를 낸다 — negative_chunk_ids가
    채워지면 all_pos의 정의(§ review_checklist.md 2-1)와 모순된다."""
    if data_type == "all_pos":
        return []
    return [
        cc.chunk for cc in candidate_pool
        if not cc.is_positive and cc.neg_grade == "hard"
    ][:limit]


def build_qa_df(items: list[QAItem], as_of_date: str | None = None) -> pd.DataFrame:
    as_of_date = as_of_date or format_kr_date(datetime.now())
    rows = []
    for it in items:
        neg_chunks = _hard_negative_chunks(it.candidate_pool, it.data_type)
        rows.append({
            "index": item_id(it.chunking_method, it.seq),
            "as_of_date": as_of_date,
            "chunking_method": it.chunking_method,
            "data_type": it.data_type,
            "difficulty": it.difficulty,
            "persona": it.persona,
            "category": it.category,
            "subcategory": it.subcategory,
            "query": it.query,
            "positive_docs": json.dumps(it.positive_docs, ensure_ascii=False),
            "positive_chunk_ids": json.dumps(it.positive_chunk_ids, ensure_ascii=False),
            "positive_chunk_meta": json.dumps(it.positive_chunk_meta, ensure_ascii=False),
            "negative_docs": json.dumps({c.doc_id: c.doc_nm for c in neg_chunks}, ensure_ascii=False),
            "negative_chunk_ids": json.dumps([c.chunk_id for c in neg_chunks], ensure_ascii=False),
            "negative_chunk_meta": json.dumps({c.chunk_id: c.content for c in neg_chunks}, ensure_ascii=False),
            "ground_truth": it.ground_truth,
            "nuggets": json.dumps([vars(n) for n in it.nuggets], ensure_ascii=False),
            "reason": it.reason,
            "검수": REVIEW_STATUS_DEFAULT,
            "납품일자": None,
            "폐기 사유": None,
            "query_수정": None,
        })
    return pd.DataFrame(rows, columns=QA_COLUMNS)


def build_chunk_df(items: list[QAItem]) -> pd.DataFrame:
    rows = []
    for it in items:
        for cc in it.candidate_pool:
            ch = cc.chunk
            rows.append({
                "index": item_id(it.chunking_method, it.seq),
                "chunking_method": it.chunking_method,
                "chunk_id": ch.chunk_id,
                "doc_id(uuid)": ch.doc_id,
                "doc_nm": ch.doc_nm,
                "category": ch.category,
                "subcategory": ch.subcategory,
                "is_positive": cc.is_positive,
                "neg_grade": cc.neg_grade,
                "content": ch.content,
            })
    return pd.DataFrame(rows, columns=CHUNK_COLUMNS)


def build_doc_df(chunk_df: pd.DataFrame, system_nm_lookup: dict) -> pd.DataFrame:
    if chunk_df.empty:
        return pd.DataFrame(columns=DOC_COLUMNS)
    g = (
        chunk_df.groupby(["index", "doc_id(uuid)", "doc_nm", "category", "subcategory"])["is_positive"]
        .any()
        .reset_index()
        .rename(columns={"is_positive": "is_positive_doc"})
    )
    g["system_nm"] = g["doc_id(uuid)"].map(lambda d: system_nm_lookup.get(d, "외부"))
    return g[DOC_COLUMNS]


def build_input_df(chunks: list[ChunkRef]) -> pd.DataFrame:
    """해당 청킹방식의 원본 input 파일(select_recursive_*_chunk_span.xlsx) 전체를
    그대로 담는다 — QA에 실제로 쓰였는지 여부와 무관하게 모든 문서의 모든 청크."""
    rows = [
        {
            "chunk_id": c.chunk_id,
            "uuid": c.doc_id,
            "doc_nm": c.doc_nm,
            "category": c.category,
            "subcategory": c.subcategory,
            "content": c.content,
            "topic_list": c.topic_list,
            "start_char": c.start_char,
            "end_char": c.end_char,
        }
        for c in chunks
    ]
    return pd.DataFrame(rows, columns=INPUT_COLUMNS)


def build_retrieval_records(items: list[QAItem]) -> list[dict]:
    """추론(검색성능) 평가용 레코드. candidate_pool 전체(정답+잡음)를 후보로 주고,
    top_k는 고정값(config)이라 출력 레코드에는 포함하지 않는다.

    data_type은 문항 생성 시점 값(it.data_type) 그대로다 — sims(모의 검색 재분류
    결과)를 쓰지 않는다. 이 레코드의 candidate_chunks는 candidate_pool 전체이고
    positive_chunk_ids는 GOLD 전체이므로, data_type도 그 둘의 실제 관계를 따라야
    한다. sim.data_type은 "키워드 겹침 모의 검색의 top_k 안에서 실제로 얼마나 잘
    풀리는지"를 반영한 재분류값이라 top_k 부분집합(topk_chunks)에만 의미가 있다 —
    여기 쓰면 candidate_chunks 안에 positive_chunk_ids가 그대로 들어있는데 data_type만
    "all_neg"라고 말하는 모순이 생긴다. build_answer_records의 context는 정확히
    topk_chunks로 구성되므로 거기서는 sim.data_type을 그대로 쓴다.

    positive_chunk_ids는 GOLD 정답 청크 id만 담는다(task2_eval_io_spec.md §6-②,
    §13-2 — Recall@K/Precision@K/MRR 계산의 P). sim 재분류와 무관하게 항상
    it.positive_chunk_ids 그대로다."""
    records = []
    for it in items:
        # neg_grade는 여기 넣지 않는다 — positive만 None이라 candidate_chunks만 봐도
        # 정답이 그대로 드러난다(검색 시스템에 실제로 넘어가는 파일이라 답 유출 금지).
        # neg_grade는 내부 검수용인 QA.xlsx의 negative_chunk_ids/negative_chunk_meta
        # 컬럼(hard 등급만, 최대 5개)에서만 확인한다.
        candidate_chunks = [
            {"chunk_id": cc.chunk.chunk_id, "title": cc.chunk.doc_nm, "content": cc.chunk.content}
            for cc in it.candidate_pool
        ]
        records.append({
            "item_id": item_id(it.chunking_method, it.seq),
            "data_type": it.data_type,
            "chunking_method": it.chunking_method,
            "business_domain": [it.category],
            "persona": it.persona,
            "difficulty": DIFFICULTY_EN.get(it.difficulty, it.difficulty),
            "question": it.query,
            "positive_chunk_ids": it.positive_chunk_ids,
            "candidate_chunks": candidate_chunks,
        })
    return records


def build_answer_records(items: list[QAItem], sims: dict[str, RetrievalSim]) -> list[dict]:
    """추론(답변성능) 평가용 레코드. context는 sims의 재분류된 data_type에 따라
    구성한다: all_neg는 빈 배열, all_pos/pos_neg는 모의 검색의 top_k 결과
    (topk_chunks) 그대로 — GOLD positive_chunk_ids 전체가 아니라, 실제로 top_k
    안에 뽑힌 positive/neg만큼만 담긴다(그래서 context 길이가 항상 top_k와 같다)."""
    records = []
    for it in items:
        sim = sims[it.index]
        if sim.data_type == "all_neg":
            context = []
        else:  # all_pos / pos_neg
            context = [
                {"chunk_id": cc.chunk.chunk_id, "title": cc.chunk.doc_nm, "content": cc.chunk.content}
                for cc in sim.topk_chunks
            ]
        records.append({
            "item_id": item_id(it.chunking_method, it.seq),
            "data_type": sim.data_type,
            "chunking_method": it.chunking_method,
            "business_domain": [it.category],
            "persona": it.persona,
            "difficulty": DIFFICULTY_EN.get(it.difficulty, it.difficulty),
            "question": it.query,
            "context": context,
            "nuggets": [vars(n) for n in it.nuggets],
        })
    return records


def save_json_items(records: list[dict], out_dir: str) -> None:
    """retrieval/answer 레코드를 한 파일에 몰아 담지 않고, 항목(item_id)당 파일 1개로
    쪼개 저장한다. 청킹방식 3개가 모두 같은 out_dir에 쓰더라도 item_id가 방식마다
    다른 1000단위 번호 대역(ID_BASE)을 쓰므로 파일명이 겹치지 않는다."""
    os.makedirs(out_dir, exist_ok=True)
    for r in records:
        with open(f"{out_dir}/{r['item_id']}.json", "w", encoding="utf-8") as f:
            json.dump(r, f, ensure_ascii=False, indent=2)


_DIFFICULTY_DISPLAY_ORDER = ["하", "중", "상", "easy", "normal", "hard"]


def save_report_xlsx(items_by_method: dict[str, list[QAItem]], out_path: str) -> None:
    """청킹방식별로 data_type×난이도 개수, persona별 개수를 집계해 report.xlsx로 저장한다.
    청킹방식마다 최종 data_type이 재분류될 수 있어(§ retrieval_sim) 방식별 시트로 나눈다."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for method, items in items_by_method.items():
        ws = wb.create_sheet(method[:31])
        ws.append([f"{method} - 총 {len(items)}건"])
        ws.append([])

        ws.append(["data_type × 난이도별 개수"])
        if items:
            difficulties = sorted(
                {it.difficulty for it in items},
                key=lambda d: _DIFFICULTY_DISPLAY_ORDER.index(d) if d in _DIFFICULTY_DISPLAY_ORDER else 999,
            )
            data_types = sorted({it.data_type for it in items})
            counts: dict[tuple[str, str], int] = {}
            for it in items:
                key = (it.data_type, it.difficulty)
                counts[key] = counts.get(key, 0) + 1
            ws.append(["data_type", *difficulties, "합계"])
            for dt in data_types:
                row = [counts.get((dt, d), 0) for d in difficulties]
                ws.append([dt, *row, sum(row)])
            total_row = [sum(counts.get((dt, d), 0) for dt in data_types) for d in difficulties]
            ws.append(["합계", *total_row, sum(total_row)])
        else:
            ws.append(["(문항 없음)"])
        ws.append([])

        ws.append(["persona별 개수"])
        if items:
            persona_counts: dict[str, int] = {}
            for it in items:
                persona_counts[it.persona] = persona_counts.get(it.persona, 0) + 1
            ws.append(["persona", "문항 수"])
            for persona, n in sorted(persona_counts.items(), key=lambda kv: -kv[1]):
                ws.append([persona, n])
        else:
            ws.append(["(문항 없음)"])

    wb.save(out_path)


def _cell_value(v):
    """openpyxl은 리스트/딕셔너리 값을 셀에 그대로 못 쓴다(pandas.to_excel은 str()로
    암묵 변환해줬는데, append 방식으로 바꾸면서 그 변환을 직접 해줘야 함)."""
    return str(v) if isinstance(v, (list, dict)) else v


def _sorted_all_items(items_by_method: dict[str, list[QAItem]]) -> list[QAItem]:
    """청킹방식 3개의 items를 item_id(=ID_BASE+seq) 오름차순 하나의 리스트로 합친다.
    통합 산출물(merged qa, retrieval/answer summary)이 전부 이 순서를 공유한다."""
    all_items = [it for items in items_by_method.values() for it in items]
    all_items.sort(key=lambda it: ID_BASE[it.chunking_method] + it.seq)
    return all_items


def save_merged_qa_xlsx(
    items_by_method: dict[str, list[QAItem]], system_nm_lookup: dict, out_path: str,
    as_of_date: str | None = None,
) -> None:
    """청킹방식 3개의 QA/USED_DOCS 시트를 item_id(index) 오름차순으로 하나의 워크북에 합쳐
    저장한다. item_id는 방식별로 겹치지 않는 1000단위 대역을 쓰므로(ID_BASE) 그대로
    이어붙여도 충돌하지 않는다 — 검수자가 세 방식을 파일 하나로 훑어볼 때 쓴다."""
    from openpyxl import Workbook

    all_items = _sorted_all_items(items_by_method)

    qa_df = build_qa_df(all_items, as_of_date)
    chunk_df = build_chunk_df(all_items)
    doc_df = build_doc_df(chunk_df, system_nm_lookup)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("QA")
    ws.append(QA_COLUMNS)
    for row in qa_df.itertuples(index=False):
        ws.append([_cell_value(v) for v in row])

    ws_docs = wb.create_sheet("USED_DOCS")
    ws_docs.append(DOC_COLUMNS)
    for row in doc_df.itertuples(index=False):
        ws_docs.append([_cell_value(v) for v in row])

    wb.save(out_path)


def save_input_chunks_xlsx(chunk_pool_by_method: dict[str, list[ChunkRef]], out_path: str) -> None:
    """청킹방식별 원본 input 파일(select_recursive_*_chunk_span.xlsx) 전체를 방식당
    시트 하나로 묶어 저장한다. items와 무관하게 항상 같은 내용이라 최종 저장 1회만
    호출한다 — 예전에는 output_{method}.xlsx의 INPUT_DATA 시트가 이 역할을 했다."""
    with pd.ExcelWriter(out_path) as writer:
        for method, chunks in chunk_pool_by_method.items():
            build_input_df(chunks).to_excel(writer, sheet_name=method[:31], index=False)


RETRIEVAL_SUMMARY_COLUMNS = [
    "item_id", "data_type", "chunking_method", "business_domain", "persona", "difficulty", "question",
    "positive_chunk_ids", "positive_chunk_개수", "후보문서_개수", "후보청크_개수",
]
ANSWER_SUMMARY_COLUMNS = [
    "item_id", "data_type", "chunking_method", "business_domain", "persona", "difficulty", "question",
    "context", "context_총개수", "context_정답청크_개수", "context_잡음청크_개수", "nuggets", "nugget_개수",
]


def build_retrieval_summary_df(items_by_method: dict[str, list[QAItem]]) -> pd.DataFrame:
    """retrieval.json(문항당 candidate_chunks 최대 ~2,000개)을 문항 1행짜리 요약으로
    바꾼다 — candidate_chunks 원문은 담지 않고, 검수자가 훑어볼 문서/청크 개수만 집계한다."""
    rows = []
    for it in _sorted_all_items(items_by_method):
        doc_ids = {cc.chunk.doc_id for cc in it.candidate_pool}
        rows.append({
            "item_id": item_id(it.chunking_method, it.seq),
            "data_type": it.data_type,
            "chunking_method": it.chunking_method,
            "business_domain": it.category,
            "persona": it.persona,
            "difficulty": DIFFICULTY_EN.get(it.difficulty, it.difficulty),
            "question": it.query,
            "positive_chunk_ids": json.dumps(it.positive_chunk_ids, ensure_ascii=False),
            "positive_chunk_개수": len(it.positive_chunk_ids),
            "후보문서_개수": len(doc_ids),
            "후보청크_개수": len(it.candidate_pool),
        })
    return pd.DataFrame(rows, columns=RETRIEVAL_SUMMARY_COLUMNS)


def save_retrieval_summary_xlsx(items_by_method: dict[str, list[QAItem]], out_path: str) -> None:
    build_retrieval_summary_df(items_by_method).to_excel(out_path, index=False, sheet_name="retrieval")


def build_answer_summary_df(
    items_by_method: dict[str, list[QAItem]], sims_by_method: dict[str, dict[str, RetrievalSim]],
) -> pd.DataFrame:
    """answer.json을 문항 1행짜리 요약으로 바꾼다. context(모의 검색 top_k) 원문은 JSON
    문자열로 담되, 그 안의 실제 positive/잡음 개수는 sim.topk_chunks의 is_positive를
    그대로 세어 별도 컬럼으로 바로 보이게 한다(문자열 비교 없이 타입 그대로 집계)."""
    rows = []
    for it in _sorted_all_items(items_by_method):
        sim = sims_by_method[it.chunking_method][it.index]
        context_chunks = [] if sim.data_type == "all_neg" else sim.topk_chunks
        n_pos = sum(1 for cc in context_chunks if cc.is_positive)
        context_json = [
            {"chunk_id": cc.chunk.chunk_id, "title": cc.chunk.doc_nm, "content": cc.chunk.content}
            for cc in context_chunks
        ]
        rows.append({
            "item_id": item_id(it.chunking_method, it.seq),
            "data_type": sim.data_type,
            "chunking_method": it.chunking_method,
            "business_domain": it.category,
            "persona": it.persona,
            "difficulty": DIFFICULTY_EN.get(it.difficulty, it.difficulty),
            "question": it.query,
            "context": json.dumps(context_json, ensure_ascii=False),
            "context_총개수": len(context_chunks),
            "context_정답청크_개수": n_pos,
            "context_잡음청크_개수": len(context_chunks) - n_pos,
            "nuggets": json.dumps([vars(n) for n in it.nuggets], ensure_ascii=False),
            "nugget_개수": len(it.nuggets),
        })
    return pd.DataFrame(rows, columns=ANSWER_SUMMARY_COLUMNS)


def save_answer_summary_xlsx(
    items_by_method: dict[str, list[QAItem]], sims_by_method: dict[str, dict[str, RetrievalSim]], out_path: str,
) -> None:
    build_answer_summary_df(items_by_method, sims_by_method).to_excel(out_path, index=False, sheet_name="answer")
